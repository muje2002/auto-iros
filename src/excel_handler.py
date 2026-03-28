"""엑셀 파일 기반 주소 일괄 입력 처리"""

import os
from openpyxl import Workbook, load_workbook

from .codef_api import (
    RegisterRequest,
    REALTY_TYPES,
    ISSUE_TYPES,
    INQUIRY_TYPES,
    REGISTER_TYPES,
)


# 엑셀 컬럼 헤더 정의
HEADERS = [
    "조회구분",
    "주소",
    "동",
    "호",
    "부동산구분",
    "등기유형",
    "발급유형",
    "고유번호",
    "시도",
    "읍면동",
    "지번",
    "시군구",
    "도로명",
    "건물번호",
]

# 기본값
DEFAULTS = {
    "조회구분": "간편검색",
    "부동산구분": "토지+건물",
    "등기유형": "전체",
    "발급유형": "열람",
}


def create_template(output_path: str) -> str:
    """입력용 엑셀 템플릿 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "등기부등본 요청"

    # 헤더 작성
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = cell.font.copy(bold=True)

    # 컬럼 너비 조정
    widths = [12, 50, 8, 8, 14, 10, 10, 20, 12, 12, 10, 12, 15, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    # 안내 시트
    guide = wb.create_sheet("입력안내")
    guide_data = [
        ["컬럼", "설명", "필수여부", "허용값"],
        ["조회구분", "조회 방식", "선택(기본:간편검색)", ", ".join(INQUIRY_TYPES.keys())],
        ["주소", "검색어 (간편검색 시)", "조건부", "최소 3자리"],
        ["동", "집합건물의 동", "선택", ""],
        ["호", "집합건물의 호", "선택", ""],
        ["부동산구분", "부동산 유형", "선택(기본:토지+건물)", ", ".join(REALTY_TYPES.keys())],
        ["등기유형", "등기 조회 범위", "선택(기본:전체)", ", ".join(REGISTER_TYPES.keys())],
        ["발급유형", "열람 또는 발급", "선택(기본:열람)", ", ".join(ISSUE_TYPES.keys())],
        ["고유번호", "14자리 (고유번호 조회 시)", "조건부", "0000-0000-000000"],
        ["시도", "시/도 (소재지번, 도로명 시)", "조건부", ""],
        ["읍면동", "읍면동/리 (소재지번 시)", "조건부", ""],
        ["지번", "지번 (소재지번 시)", "조건부", ""],
        ["시군구", "시군구 (도로명 시)", "조건부", ""],
        ["도로명", "도로명 (도로명 시)", "조건부", ""],
        ["건물번호", "건물번호 (도로명 시)", "조건부", ""],
    ]
    for row_idx, row_data in enumerate(guide_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            guide.cell(row=row_idx, column=col_idx, value=value)

    # 예시 데이터
    examples = [
        ["간편검색", "테헤란로 123", "", "", "토지+건물", "전체", "열람",
         "", "", "", "", "", "", ""],
        ["간편검색", "서초대로 456", "101동", "202호", "집합건물", "전체", "열람",
         "", "", "", "", "", "", ""],
        ["고유번호", "", "", "", "", "전체", "열람",
         "1101-2024-123456", "", "", "", "", "", ""],
    ]
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)
    return output_path


def read_requests(file_path: str) -> list[RegisterRequest]:
    """엑셀 파일에서 등기부등본 요청 목록 읽기"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    requests_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue

        def cell_val(idx: int) -> str:
            if idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ""

        inquiry_type = cell_val(0) or DEFAULTS["조회구분"]
        address = cell_val(1)
        unique_no = cell_val(7)

        # 최소 검증: 간편검색이면 주소, 고유번호면 고유번호 필요
        if inquiry_type == "간편검색" and not address:
            continue
        if inquiry_type == "고유번호" and not unique_no:
            continue

        req = RegisterRequest(
            inquiry_type=inquiry_type,
            address=address,
            dong=cell_val(2),
            ho=cell_val(3),
            realty_type=cell_val(4) or DEFAULTS["부동산구분"],
            register_type=cell_val(5) or DEFAULTS["등기유형"],
            issue_type=cell_val(6) or DEFAULTS["발급유형"],
            unique_no=unique_no,
            addr_sido=cell_val(8),
            addr_dong=cell_val(9),
            addr_lot_number=cell_val(10),
            addr_sigungu=cell_val(11),
            addr_road_name=cell_val(12),
            addr_building_number=cell_val(13),
        )
        requests_list.append(req)

    wb.close()

    if not requests_list:
        raise ValueError(f"엑셀 파일에 유효한 요청 데이터가 없습니다: {file_path}")

    return requests_list


def export_results(summaries: list[dict], output_path: str) -> str:
    """배치 조회 결과를 엑셀 파일로 내보내기.

    Args:
        summaries: save_batch_pdfs()가 반환한 결과 요약 리스트
        output_path: 저장할 엑셀 파일 경로

    Returns:
        저장된 파일 경로
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "조회 결과"

    # 헤더
    result_headers = ["#", "주소/고유번호", "상태", "PDF 파일", "오류"]
    for col, header in enumerate(result_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = cell.font.copy(bold=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40

    # 데이터
    for i, s in enumerate(summaries, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=s.get("address", ""))
        ws.cell(row=i + 1, column=3, value=s.get("status", ""))
        ws.cell(row=i + 1, column=4, value=s.get("file") or "")
        ws.cell(row=i + 1, column=5, value=s.get("error") or "")

    # 요약 시트
    summary_sheet = wb.create_sheet("요약")
    success_count = sum(1 for s in summaries if s.get("status") == "성공")
    fail_count = len(summaries) - success_count
    summary_data: list[list[str | int]] = [
        ["항목", "값"],
        ["총 건수", len(summaries)],
        ["성공", success_count],
        ["실패", fail_count],
    ]
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            summary_sheet.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)
    return output_path
