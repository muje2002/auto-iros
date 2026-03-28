"""excel_handler.py 단위테스트"""

import os
import tempfile

from openpyxl import load_workbook

from src.excel_handler import create_template, export_results, read_requests


class TestCreateTemplate:
    def test_creates_file(self):
        tmp = os.path.join(tempfile.mkdtemp(), "template.xlsx")
        path = create_template(tmp)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 1000
        os.unlink(path)

    def test_has_correct_sheets(self):
        tmp = os.path.join(tempfile.mkdtemp(), "template.xlsx")
        create_template(tmp)
        wb = load_workbook(tmp, read_only=True)
        assert "등기부등본 요청" in wb.sheetnames
        assert "입력안내" in wb.sheetnames
        wb.close()
        os.unlink(tmp)

    def test_has_example_rows(self):
        tmp = os.path.join(tempfile.mkdtemp(), "template.xlsx")
        create_template(tmp)
        wb = load_workbook(tmp, read_only=True)
        ws = wb["등기부등본 요청"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 2  # 예시 데이터 최소 2행
        wb.close()
        os.unlink(tmp)


class TestReadRequests:
    def test_read_from_template(self):
        """생성한 템플릿을 읽으면 예시 데이터가 파싱됨"""
        tmp = os.path.join(tempfile.mkdtemp(), "template.xlsx")
        create_template(tmp)
        requests_list = read_requests(tmp)
        assert len(requests_list) >= 2
        assert requests_list[0].inquiry_type == "간편검색"
        os.unlink(tmp)

    def test_file_not_found(self):
        import pytest
        with pytest.raises(FileNotFoundError):
            read_requests("/nonexistent/file.xlsx")


class TestExportResults:
    def test_export_creates_file(self):
        summaries = [
            {"address": "서울 강남구", "status": "성공", "file": "/out/a.pdf"},
            {"address": "서울 서초구", "status": "실패", "file": None, "error": "오류"},
        ]
        tmp = os.path.join(tempfile.mkdtemp(), "result.xlsx")
        path = export_results(summaries, tmp)
        assert os.path.exists(path)

        wb = load_workbook(path, read_only=True)
        assert "조회 결과" in wb.sheetnames
        assert "요약" in wb.sheetnames

        ws = wb["요약"]
        assert ws.cell(2, 2).value == 2  # 총 건수
        assert ws.cell(3, 2).value == 1  # 성공
        assert ws.cell(4, 2).value == 1  # 실패
        wb.close()
        os.unlink(path)
